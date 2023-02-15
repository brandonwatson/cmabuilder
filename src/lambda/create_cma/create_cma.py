# necessary imports
import os
import boto3
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors, Border, Side
from openpyxl.utils import cell
# from tempfile import NamedTemporaryFile

# any global vars


def fetch_data_from_JSON():
    f = open('test_data.json')
    return json.load(f)


def insert_metadata():
    meta = cmadata['doc metadata']
    ws['A1'] = meta['realtor name'] + " CMA"
    ws['A2'] = "Created by: " + meta['generated by']


def get_col_num(col_name):
    col_num = -1
    for k, v in cmadata['headers'].items():
        if v == col_name:
            col_num = int(k)

    if col_num == -1:
        for k, v in cmadata['comparison headers'].items():
            if v == col_name:
                col_num = int(k)

    # this will return 0 if col not found, and a valid col if found
    return col_num + 1


def create_property_table(cur_col, cur_row, propertytype):
    # generate the table headers
    for row in ws.iter_rows(min_row=cur_row, max_row=cur_row, min_col=cur_col, max_col=cur_col+len(cmadata['headers'])-1):
        i = 0
        for cell in row:
            cell.value = cmadata['headers'].get(str(i))

            font_styles = cmadata['style data']['header']
            cell.font = Font(
                name=font_styles['font name'], size=font_styles['font size'], bold=font_styles['font bold'])
            i += 1

    # generate the table data
    # TODO: Need to make sure that when the JSON is written, the values for Above Grade and Effective SqFt are calculated
    # and inserted into the JSON based on the values give in UX for main sqft, upper sqft, finished basement, and unfinished
    # basement.
    cur_row += 1
    for property in cmadata[propertytype]:
        for row in ws.iter_rows(min_row=cur_row, max_row=cur_row, min_col=cur_col, max_col=cur_col + len(cmadata['headers'])-1):
            i = 0
            for activecell in row:
                font_styles = cmadata['style data'][cmadata['formats']['property']
                                                    [cmadata['headers'].get(str(i))]]
                activecell.value = property.get(
                    cmadata['headers'].get(str(i)))

                activecell.font = Font(
                    name=font_styles['font name'], size=font_styles['font size'], bold=font_styles['font bold'])
                activecell.number_format = font_styles['format']
                i += 1
        cur_row += 1


def create_property_table_calcs(cur_row, cur_col, propertytype):
    # using the forumla key allows me to have the formula ready to go and the tokens in place
    # for the values like: sale_col, active_row, list_col, active_row

    sale_price_col = get_col_num("Sale Price")
    list_price_col = get_col_num("List Price")

    if (sale_price_col or list_price_col) == 0:
        print("ERROR: could not get {0} or {1}", ("Sale Price", "List Price"))

    i = 0

    for k, v in cmadata['property calculations'].items():
        active_row = cur_row
        active_col = cur_col + \
            len(cmadata['headers']) + \
            i  # i is to track the insertions of caluclation columns

        numerator = get_col_num(v['numerator'])
        denominator = get_col_num(v['denominator'])
        if (numerator or denominator) == 0:
            print("ERROR: could not get {0} or {1}",
                  (v['numerator']), v['denominator'])
            break

        activecell = ws.cell(row=active_row,
                             column=active_col)
        activecell.value = v['label']

        # this is the header rows
        font_styles = cmadata['style data']['header']
        activecell.font = Font(
            name=font_styles['font name'], size=font_styles['font size'], bold=font_styles['font bold'])
        activecell.number_format = font_styles['format']
        active_row += 1

        for property in cmadata[propertytype]:
            numerator_col = cell._get_column_letter(numerator)
            demoninator_col = cell._get_column_letter(denominator)

            # now generate the formula
            formula = v['formula'].format(
                numerator_col, active_row, demoninator_col, active_row)
            activecell = ws.cell(row=active_row, column=active_col)
            activecell.value = formula

            font_styles = cmadata['style data'][cmadata['formats']
                                                ['property'][k]]
            activecell.font = Font(
                name=font_styles['font name'], size=font_styles['font size'], bold=font_styles['font bold'])
            activecell.number_format = font_styles['format']
            active_row += 1

        i += 1

    return i


def create_comparison_data():
    comparisons = cmadata['comparison table']['tables']
    sqft_use = cmadata['comparison table']['sqft use']

    dataObj = {}

    for compare_k, compare_v in comparisons.items():
        for sqft_v in sqft_use.values():
            dollar_sqft_comps_list = [x['Sale Price'] /
                                      x[sqft_v] for x in cmadata['comparables']]

            if compare_v['formulas']['drop minimum'] == True:
                dollar_sqft_comps_list.remove(min(dollar_sqft_comps_list))

            if compare_v['formulas']['drop maximum'] == True:
                dollar_sqft_comps_list.remove(max(dollar_sqft_comps_list))

            offer_price = cmadata['listing'][0]['Sale Price']
            sqft = cmadata['listing'][0][sqft_v]
            avg_dollar_sqft_comps = sum(
                dollar_sqft_comps_list) / len(dollar_sqft_comps_list)
            implied_price_comps = sqft * avg_dollar_sqft_comps
            implied_dollar_sqft = offer_price / sqft
            diff_dollar_sqft = (offer_price / sqft) - avg_dollar_sqft_comps
            diff_percent_sqft = ((offer_price / sqft) -
                                 avg_dollar_sqft_comps) / (avg_dollar_sqft_comps)

            if compare_k not in dataObj:
                dataObj[compare_k] = {}

            dataObj[compare_k][sqft_v] = {
                # "offer_price": offer_price,
                # "sqft": sqft,
                # "avg_dollar_sqft_comps": avg_dollar_sqft_comps,
                "Implied Price Comps": implied_price_comps,
                "Implied Dollar SqFt Offer Price": implied_dollar_sqft,
                "Dollar Diff SqFt": diff_dollar_sqft,
                "Percent Diff SqFt": diff_percent_sqft
            }

    return dataObj


def create_comparison_tables(data):
    anchors = cmadata['anchors']['comparisons']
    anchor_row = anchors['row']
    anchor_col = len(cmadata['headers']) + max(cmadata['anchors']['listing']
                                               ['formulas inserted'], cmadata['anchors']['comparables']['formulas inserted']) + 1 + anchors['col buffer']

    table_rows_offset = 1 + \
        len(cmadata['formats']['comparisons']) + anchors['row buffer']

    for num_tables, (data_k, data_v) in enumerate(data.items(), start=0):
        active_row = anchor_row + (num_tables * table_rows_offset)

        # write the name of the table
        activecell = ws[f"{cell._get_column_letter(anchor_col)}{active_row}"]
        font_styles = cmadata['style data']['header']
        activecell.value = data_k
        activecell.font = Font(
            name=font_styles['font name'], size=font_styles['font size'], bold=font_styles['font bold'])
        activecell.number_format = font_styles['format']

        for col_offset, (header_k, header_v) in enumerate(cmadata['comparison table']['sqft use'].items(), start=1):
            activecell = ws[f"{cell._get_column_letter(anchor_col + col_offset)}{active_row}"]
            activecell.value = header_v
            activecell.font = Font(
                name=font_styles['font name'], size=font_styles['font size'], bold=font_styles['font bold'])
            activecell.number_format = font_styles['format']

        # now write the row lables
        active_row += 1
        for row_offset, (label_k, label_format_v) in enumerate(cmadata['formats']['comparisons'].items(), start=0):
            activecell = ws[f"{cell._get_column_letter(anchor_col)}{active_row+row_offset}"]
            activecell.value = label_k
            activecell.font = Font(
                name=font_styles['font name'], size=font_styles['font size'], bold=font_styles['font bold'])
            activecell.number_format = font_styles['format']

        # now write the values of the cols based on the row lanels
        for col_offset, (header_k, header_v) in enumerate(cmadata['comparison table']['sqft use'].items(), start=1):
            rows_written = 0
            label_val = ws[f"{cell._get_column_letter(anchor_col)}{active_row + rows_written}"].value
            while (label_val != None):
                activecell = ws[f"{cell._get_column_letter(anchor_col + col_offset)}{active_row + rows_written}"]
                activecell.value = data_v[header_v][label_val]
                font_styles = cmadata['style data'][cmadata['formats']
                                                    ['comparisons'][label_val]]
                activecell.font = Font(
                    name=font_styles['font name'], size=font_styles['font size'], bold=font_styles['font bold'])
                activecell.number_format = font_styles['format']
                rows_written += 1
                label_val = ws[f"{cell._get_column_letter(anchor_col)}{active_row+rows_written}"].value


        # MAIN CODE
        # set the working directory for this lambda
os.chdir('./src/lambda/create_cma')

# Using an Excel workbook file extension
wb = Workbook()
ws = wb.active

# this is a simluated call to the server to get the JSON we need to run the code
cmadata = fetch_data_from_JSON()

insert_metadata()

# get the anchor data for the listing tables
anchor_dict = cmadata['anchors'].copy()

# need to remove some data from JSON - may be a better way to store this data in JSON but for now this worls
remove_keys = ['comparisons']
for x in remove_keys:
    del anchor_dict[x]

# create the listing and property tables
for k, v in anchor_dict.items():
    create_property_table(cur_col=v['col'],
                          cur_row=v['row'],
                          propertytype=v['propertytype'])

    cmadata['anchors'][k]['formulas inserted'] += create_property_table_calcs(cur_row=v['row'],
                                                                              cur_col=v['col'],
                                                                              propertytype=v['propertytype'])

# now create the comparison tables
create_comparison_tables(create_comparison_data())

# begin: ecessary small change to a cell value
font_styles = cmadata['style data']['header']

ws.cell(row=cmadata['anchors']['listing']['row'], column=cmadata['anchors']['listing']['col'] +
        len(cmadata['headers']) - 1).value = "Offer Price"

ws.cell(row=cmadata['anchors']['listing']['row'], column=cmadata['anchors']['listing']['col'] +
        len(cmadata['headers']) - 1).font = Font(
    name=font_styles['font name'], size=font_styles['font size'], bold=font_styles['font bold'])
# end: necessary small change


# ... Build out workbook with openpyxl
# s3_resource = boto3.resource('s3')
dest_filename = "my-file.xlsx"

# don't forget on windows we are using backslash
filename = '.\\tmp\\{}'.format(dest_filename)
wb.save(filename)

# This is for saving into Lambda environment
# with NamedTemporaryFile() as tmp:
#    filename = '/tmp/{}'.format(dest_filename)
#    wb.save(filename)
# s3_resource.Bucket(bucket_name).upload_file(Filename=filename, Key=dest_filename)
# create downloadable URL (see other tutorial)
# https://danh-was-here.netlify.app/download-aws-s3-files
# url = create_presigned_url(bucket_name, dest_filename)
# return {
#    'downloadurl': url
# }
